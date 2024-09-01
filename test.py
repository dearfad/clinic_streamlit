from openpyxl import Workbook
import json

wb = Workbook()
ws = wb.active
ws.title = "my_data"

with open("data/patients.json", "r", encoding="utf-8") as file:
        data = json.load(file)

d = data[0]['info']

column = 1
for key, value in d.items():
    ws.cell(row=1, column=column, value=key)
    ws.cell(row=2, column=column, value=value)
    column += 1

wb.save('d.xlsx')